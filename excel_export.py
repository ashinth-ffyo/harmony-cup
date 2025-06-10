import openpyxl
from openpyxl.styles import Font

def export_to_excel(db_manager, model):
    """Export all team data to an Excel file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for category in model.categories:
        ws = wb.create_sheet(category)
        ws.append(model.columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        teams = db_manager.get_teams(category)
        for team in teams:
            ws.append([team.get(col, "") for col in model.columns])

    wb.save("Harmony Cup.xlsx")